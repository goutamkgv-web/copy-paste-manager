import csv
import ctypes
import os
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
except ImportError:
    openpyxl = None

APP_TITLE = "Copy-Paste Manager"
DEFAULT_DELAY = 1500
CLIPBOARD_RETRY_COUNT = 40
CLIPBOARD_RETRY_SECONDS = 0.05
CLIPBOARD_SETTLE_SECONDS = 0.25

user32 = ctypes.windll.user32
kernel32 = ctypes.windll.kernel32
CF_UNICODETEXT = 13
GMEM_MOVEABLE = 0x0002

# Explicit Win32 signatures are important on 64-bit Windows. Without them,
# ctypes can treat HGLOBAL/HANDLE values as 32-bit integers and truncate the
# pointer returned by GlobalAlloc, causing GlobalLock to fail.
kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
kernel32.GlobalAlloc.restype = ctypes.c_void_p
kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
kernel32.GlobalLock.restype = ctypes.c_void_p
kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
kernel32.GlobalUnlock.restype = ctypes.c_int
kernel32.GlobalFree.argtypes = [ctypes.c_void_p]
kernel32.GlobalFree.restype = ctypes.c_void_p

user32.OpenClipboard.argtypes = [ctypes.c_void_p]
user32.OpenClipboard.restype = ctypes.c_int
user32.EmptyClipboard.argtypes = []
user32.EmptyClipboard.restype = ctypes.c_int
user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
user32.SetClipboardData.restype = ctypes.c_void_p
user32.CloseClipboard.argtypes = []
user32.CloseClipboard.restype = ctypes.c_int
user32.GetClipboardData.argtypes = [ctypes.c_uint]
user32.GetClipboardData.restype = ctypes.c_void_p


def _read_clipboard_text():
    """Read back CF_UNICODETEXT for verification. Returns None if unavailable."""
    if not user32.OpenClipboard(None):
        return None
    try:
        h = user32.GetClipboardData(CF_UNICODETEXT)
        if not h:
            return None
        p = kernel32.GlobalLock(h)
        if not p:
            return None
        try:
            return ctypes.wstring_at(p)
        finally:
            kernel32.GlobalUnlock(h)
    finally:
        user32.CloseClipboard()


def set_clipboard_text(text: str):
    """Set clipboard text and verify exact read-back before continuing.

    Clipboard History is asynchronous. The caller must also allow a generous
    settle interval between items. This function verifies the actual Windows
    clipboard contents so a failed SetClipboardData is never reported as a
    successful copy.
    """
    text = "" if text is None else str(text)
    data = text.encode("utf-16-le") + b"\x00\x00"
    h = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(data))
    if not h:
        raise OSError("Windows could not allocate clipboard memory.")

    p = kernel32.GlobalLock(h)
    if not p:
        kernel32.GlobalFree(h)
        raise OSError("Windows could not lock clipboard memory.")
    try:
        ctypes.memmove(p, data, len(data))
    finally:
        kernel32.GlobalUnlock(h)

    opened = False
    for _ in range(CLIPBOARD_RETRY_COUNT):
        if user32.OpenClipboard(None):
            opened = True
            break
        time.sleep(CLIPBOARD_RETRY_SECONDS)

    if not opened:
        kernel32.GlobalFree(h)
        raise OSError("Could not open the Windows clipboard after repeated retries. Another application may be using it.")

    try:
        if not user32.EmptyClipboard():
            raise OSError("Windows could not clear the clipboard.")
        if not user32.SetClipboardData(CF_UNICODETEXT, h):
            raise OSError("Windows rejected the clipboard value.")
        # Ownership transfers to Windows after SetClipboardData succeeds.
        h = None
    finally:
        user32.CloseClipboard()
        if h:
            kernel32.GlobalFree(h)

    # Verify the actual clipboard value. This protects against reporting a
    # successful copy when another application immediately replaced it.
    deadline = time.monotonic() + 1.5
    while time.monotonic() < deadline:
        actual = _read_clipboard_text()
        if actual == text:
            # Give Clipboard History a small, deterministic settle period.
            time.sleep(CLIPBOARD_SETTLE_SECONDS)
            return
        time.sleep(0.05)

    actual = _read_clipboard_text()
    raise OSError(
        "Clipboard verification failed. The clipboard was changed by another application "
        f"before verification completed (expected {len(text)} characters, got "
        f"{len(actual) if actual is not None else 'no text'})."
    )


class CopyPasteManager(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1050x850")
        self.minsize(950, 760)

        self.file_path = ""
        self.workbook = None
        self.sheet_names = []
        self.values = []
        self.status_rows = []
        self.stop_event = threading.Event()
        self.processing = False
        self.last_status_snapshot = None
        self.last_update_rows = []
        self.current_task = None

        self.column_var = tk.StringVar(value="A")
        self.from_var = tk.StringVar(value="1")
        self.to_var = tk.StringVar(value="10")
        self.status_col_var = tk.StringVar(value="C")
        self.delay_var = tk.IntVar(value=DEFAULT_DELAY)
        self.mode_var = tk.StringVar(value="Skip Completed")
        self.sheet_var = tk.StringVar()
        self.count_var = tk.StringVar(value="No values loaded.")
        self.status_var = tk.StringVar(
            value="Safe Mode ON • Select a spreadsheet to begin."
        )
        self.progress_value = tk.DoubleVar(value=0)

        self.build_ui()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass

        # Keep the primary actions outside the scrolling area so they are
        # always visible on small screens and at 125%/150% Windows scaling.
        self.geometry("1050x850")
        self.minsize(900, 650)

        outer = ttk.Frame(self, padding=14)
        outer.pack(fill="both", expand=True)

        # Fixed action bar at the bottom.
        action_bar = ttk.LabelFrame(
            outer, text="ACTIONS — ALWAYS AVAILABLE", padding=8
        )
        action_bar.pack(side="bottom", fill="x", pady=(8, 0))

        self.load_btn = ttk.Button(
            action_bar,
            text="▶  START PROCESSING",
            command=self.start
        )
        self.load_btn.pack(side="left", fill="x", expand=True, ipady=10)

        self.stop_btn = ttk.Button(
            action_bar, text="■  STOP", command=self.stop, state="disabled"
        )
        self.stop_btn.pack(side="left", padx=(8, 0), ipady=10)

        self.complete_btn = ttk.Button(
            action_bar,
            text="✓  APPROVE & MARK COMPLETED",
            command=self.request_completion,
            state="disabled"
        )
        self.complete_btn.pack(side="left", padx=(8, 0), ipady=10)

        self.reset_btn = ttk.Button(
            action_bar, text="RESET", command=self.reset_status
        )
        self.reset_btn.pack(side="left", padx=(8, 0), ipady=10)

        # Fixed progress/status area above the action bar.
        self.progress = ttk.Progressbar(
            outer, variable=self.progress_value, maximum=100
        )
        self.progress.pack(side="bottom", fill="x", pady=(5, 2))
        ttk.Label(
            outer, textvariable=self.status_var
        ).pack(side="bottom", anchor="w", pady=(0, 2))

        # Scrollable main content.
        scroll_host = ttk.Frame(outer)
        scroll_host.pack(side="top", fill="both", expand=True)

        canvas = tk.Canvas(scroll_host, highlightthickness=0, borderwidth=0)
        scrollbar = ttk.Scrollbar(
            scroll_host, orient="vertical", command=canvas.yview
        )
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        content = ttk.Frame(canvas, padding=(6, 4, 6, 10))
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")

        def update_scrollregion(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def resize_content(event):
            canvas.itemconfigure(window_id, width=event.width)

        content.bind("<Configure>", update_scrollregion)
        canvas.bind("<Configure>", resize_content)

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        top = ttk.Frame(content)
        top.pack(fill="x")
        ttk.Label(
            top, text=APP_TITLE, font=("Segoe UI", 22, "bold")
        ).pack(side="left")
        ttk.Label(
            top, text="🔒 SAFE MODE: ON",
            font=("Segoe UI", 10, "bold")
        ).pack(side="right")

        ttk.Label(
            content,
            text="Select spreadsheet → enter range → Preview → Start Processing → verify → approve completion.",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(2, 8))

        privacy = ttk.LabelFrame(content, text="Safety & privacy", padding=8)
        privacy.pack(fill="x", pady=(0, 8))
        ttk.Label(
            privacy,
            text="Local only • No upload • No cloud/API • No telemetry • No database • No startup/Registry changes"
        ).pack(anchor="w")
        ttk.Label(
            privacy,
            text="Only the selected file/range is processed. Spreadsheet writing is separate and approval-gated."
        ).pack(anchor="w", pady=(2, 0))

        file_box = ttk.LabelFrame(content, text="1. INPUT", padding=9)
        file_box.pack(fill="x", pady=(0, 8))
        self.file_label = ttk.Label(file_box, text="No file selected")
        self.file_label.pack(side="left", fill="x", expand=True)
        ttk.Button(
            file_box, text="Browse…", command=self.choose_file
        ).pack(side="right")

        controls = ttk.LabelFrame(content, text="2. DATA + STATUS", padding=9)
        controls.pack(fill="x", pady=(0, 8))

        labels = [
            ("Sheet", 0), ("Data column", 1), ("From row", 2),
            ("To row", 3), ("Status column", 4), ("Processing mode", 5),
            ("Delay (ms)", 6)
        ]
        for label, col in labels:
            ttk.Label(controls, text=label).grid(
                row=0, column=col, sticky="w"
            )

        self.sheet_combo = ttk.Combobox(
            controls, textvariable=self.sheet_var,
            state="disabled", width=19
        )
        self.sheet_combo.grid(row=1, column=0, padx=(0, 8), sticky="w")
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.preview())

        self.data_entry = ttk.Entry(
            controls, textvariable=self.column_var, width=9
        )
        self.data_entry.grid(row=1, column=1, padx=(0, 8), sticky="w")
        self.data_entry.bind("<KeyRelease>", lambda e: self.preview())

        self.from_entry = ttk.Entry(
            controls, textvariable=self.from_var, width=9
        )
        self.from_entry.grid(row=1, column=2, padx=(0, 8), sticky="w")
        self.from_entry.bind("<KeyRelease>", lambda e: self.preview())

        self.to_entry = ttk.Entry(
            controls, textvariable=self.to_var, width=9
        )
        self.to_entry.grid(row=1, column=3, padx=(0, 8), sticky="w")
        self.to_entry.bind("<KeyRelease>", lambda e: self.preview())

        self.status_entry = ttk.Entry(
            controls, textvariable=self.status_col_var, width=9
        )
        self.status_entry.grid(row=1, column=4, padx=(0, 8), sticky="w")

        self.mode_combo = ttk.Combobox(
            controls,
            textvariable=self.mode_var,
            values=["Skip Completed", "Copy All", "Pending Only", "Resume"],
            state="readonly", width=16
        )
        self.mode_combo.grid(row=1, column=5, padx=(0, 8), sticky="w")
        self.mode_combo.bind("<<ComboboxSelected>>", lambda e: self.preview())

        ttk.Spinbox(
            controls, from_=500, to=5000, increment=100,
            textvariable=self.delay_var, width=9
        ).grid(row=1, column=6, sticky="w")

        ttk.Label(
            controls,
            text="Completed recognized as: Completed / ✓ Completed"
        ).grid(row=2, column=0, columnspan=7, sticky="w", pady=(6, 0))

        check_box = ttk.LabelFrame(
            content, text="3. CHECK — DRY RUN SUMMARY", padding=8
        )
        check_box.pack(fill="x", pady=(0, 8))

        self.summary_text = tk.Text(
            check_box, height=5, wrap="word",
            font=("Segoe UI", 10), state="disabled"
        )
        self.summary_text.pack(fill="x")

        preview_box = ttk.LabelFrame(
            content, text="4. PREVIEW", padding=8
        )
        preview_box.pack(fill="x", pady=(0, 8))

        preview_header = ttk.Frame(preview_box)
        preview_header.pack(fill="x")
        ttk.Label(
            preview_header, textvariable=self.count_var,
            font=("Segoe UI", 9, "bold")
        ).pack(side="left", fill="x", expand=True)

        ttk.Button(
            preview_header, text="↻  PREVIEW / VALIDATE",
            command=self.preview
        ).pack(side="right")

        preview_frame = ttk.Frame(preview_box)
        preview_frame.pack(fill="x", pady=(5, 0))

        self.preview_text = tk.Text(
            preview_frame, height=6, wrap="none",
            font=("Consolas", 9), state="disabled"
        )
        preview_scroll = ttk.Scrollbar(
            preview_frame, orient="vertical",
            command=self.preview_text.yview
        )
        self.preview_text.configure(yscrollcommand=preview_scroll.set)
        self.preview_text.pack(side="left", fill="both", expand=True)
        preview_scroll.pack(side="right", fill="y")

        status_box = ttk.LabelFrame(
            content, text="5. RESULT — COPIED CONTENT STATUS", padding=8
        )
        status_box.pack(fill="both", expand=True)

        cols = ("row", "content", "clipboard", "sheetstatus")
        self.tree = ttk.Treeview(
            status_box, columns=cols, show="headings", height=9
        )
        self.tree.heading("row", text="Row")
        self.tree.heading("content", text="Content")
        self.tree.heading("clipboard", text="Clipboard")
        self.tree.heading("sheetstatus", text="Spreadsheet Status")
        self.tree.column("row", width=65, anchor="center")
        self.tree.column("content", width=430)
        self.tree.column("clipboard", width=110, anchor="center")
        self.tree.column("sheetstatus", width=160, anchor="center")

        scroll = ttk.Scrollbar(
            status_box, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        ttk.Label(
            content,
            text="Completion is never automatic. Spreadsheet changes require explicit approval.",
            font=("Segoe UI", 9, "italic")
        ).pack(anchor="w", pady=(7, 0))

    def set_summary(self, text):
        self.summary_text.config(state="normal")
        self.summary_text.delete("1.0", "end")
        self.summary_text.insert("end", text)
        self.summary_text.config(state="disabled")

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Select spreadsheet",
            filetypes=[
                ("Excel workbook", "*.xlsx;*.xlsm"),
                ("CSV file", "*.csv"),
            ]
        )
        if not path:
            return

        self.close_workbook()
        self.file_path = os.path.abspath(path)
        self.file_label.config(text=os.path.basename(path))

        ext = os.path.splitext(path)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            if openpyxl is None:
                messagebox.showerror(APP_TITLE, "openpyxl is missing. Run install.bat.")
                return
            try:
                self.workbook = openpyxl.load_workbook(
                    path, read_only=True, data_only=True, keep_links=False
                )
                self.sheet_names = self.workbook.sheetnames
                self.sheet_combo.config(state="readonly", values=self.sheet_names)
                self.sheet_var.set(self.sheet_names[0] if self.sheet_names else "")
            except Exception as e:
                self.close_workbook()
                messagebox.showerror(APP_TITLE, f"Could not open workbook:\n{e}")
                return
        else:
            self.sheet_names = ["CSV"]
            self.sheet_combo.config(state="disabled", values=self.sheet_names)
            self.sheet_var.set("CSV")

        self.reset_status()
        self.preview()

    def close_workbook(self):
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        self.workbook = None
        self.sheet_names = []

    def parse_column(self, raw):
        s = raw.strip().upper()
        if not s:
            raise ValueError("Column is required.")
        if s.isdigit():
            n = int(s)
            if n < 1:
                raise ValueError("Column number must be at least 1.")
            return n
        if not s.isalpha():
            raise ValueError("Column must be like A, B, C…")
        n = 0
        for ch in s:
            n = n * 26 + ord(ch) - 64
        return n

    def get_range(self):
        data_col = self.parse_column(self.column_var.get())
        status_col = self.parse_column(self.status_col_var.get())
        start = int(self.from_var.get())
        end = int(self.to_var.get())
        if start < 1 or end < start:
            raise ValueError("Enter a valid From/To row range.")
        if data_col == status_col:
            raise ValueError("Data column and Status column must be different.")
        if end - start > 100000:
            raise ValueError("Please keep a batch at or below 100,001 rows.")
        return data_col, status_col, start, end

    def read_values(self, preview_only=False):
        data_col, status_col, start, end = self.get_range()
        ext = os.path.splitext(self.file_path)[1].lower()
        result = []

        if ext == ".csv":
            with open(self.file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                for row_num, row in enumerate(reader, start=1):
                    if row_num < start:
                        continue
                    if row_num > end:
                        break
                    value = row[data_col - 1] if data_col <= len(row) else ""
                    existing = row[status_col - 1] if status_col <= len(row) else ""
                    if value is not None and str(value).strip():
                        result.append((row_num, str(value), str(existing or "")))
                    if preview_only and len(result) >= 20:
                        break
            return result

        if self.workbook is None:
            raise ValueError("Select an Excel workbook first.")
        sheet = self.sheet_var.get()
        if sheet not in self.sheet_names:
            raise ValueError("Select a worksheet.")

        ws = self.workbook[sheet]
        for row_num in range(start, end + 1):
            value = ws.cell(row=row_num, column=data_col).value
            existing = ws.cell(row=row_num, column=status_col).value
            if value is not None and str(value).strip():
                result.append((row_num, str(value), "" if existing is None else str(existing)))
            if preview_only and len(result) >= 20:
                break
        return result

    def is_completed(self, status):
        return str(status).strip().lower() in ("completed", "✓ completed")

    def calculate_task(self):
        values = self.read_values(preview_only=False)
        mode = self.mode_var.get()
        if mode == "Copy All":
            selected = values
        elif mode in ("Skip Completed", "Pending Only", "Resume"):
            selected = [x for x in values if not self.is_completed(x[2])]
        else:
            selected = values

        completed = sum(1 for x in values if self.is_completed(x[2]))
        return values, selected, completed

    def preview(self):
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")
        try:
            if not self.file_path:
                raise ValueError("No file selected.")

            values, selected, completed = self.calculate_task()
            mode = self.mode_var.get()

            summary = (
                f"File: {os.path.basename(self.file_path)}\n"
                f"Sheet: {self.sheet_var.get()}\n"
                f"Data column: {self.column_var.get().strip().upper()}    "
                f"Rows: {self.from_var.get()} → {self.to_var.get()}\n"
                f"Status column: {self.status_col_var.get().strip().upper()}    "
                f"Mode: {mode}\n"
                f"Rows containing data: {len(values):,}\n"
                f"Already Completed: {completed:,}\n"
                f"New values to process: {len(selected):,}\n"
                f"Clipboard History: Windows feature — retention is controlled by Windows.\n"
                f"Spreadsheet write: NO during copy; only after explicit completion approval."
            )
            self.set_summary(summary)

            preview_items = selected[:20]
            self.count_var.set(
                f"Preview: {len(preview_items)} of {len(selected):,} values to process"
            )
            for row_num, value, existing in preview_items:
                self.preview_text.insert("end", f"Row {row_num}: {value}\n")
            if not preview_items:
                self.preview_text.insert("end", "Nothing new to process in this mode.")
        except Exception as e:
            self.set_summary("Dry run unavailable: " + str(e))
            self.count_var.set("Preview unavailable")
            self.preview_text.insert("end", str(e))
        finally:
            self.preview_text.config(state="disabled")

    def clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def populate_status(self, values):
        self.clear_tree()
        self.status_rows = []
        for row_num, value, existing in values:
            initial = existing if existing.strip() else "Pending"
            iid = self.tree.insert(
                "", "end", values=(row_num, value, "Pending", initial)
            )
            self.status_rows.append((iid, row_num, value, existing))

    def start(self):
        if self.processing:
            return

        try:
            all_values, selected, _ = self.calculate_task()
            if not selected:
                messagebox.showinfo(
                    APP_TITLE,
                    "There are no new values to process in the selected range/mode."
                )
                return
            delay = max(500, int(self.delay_var.get()))
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        answer = messagebox.askyesno(
            "Final Dry-Run Approval",
            f"READY TO START\n\n"
            f"Rows with data: {len(all_values):,}\n"
            f"Values to copy: {len(selected):,}\n"
            f"Data column: {self.column_var.get().strip().upper()}\n"
            f"Rows: {self.from_var.get()}–{self.to_var.get()}\n"
            f"Status column: {self.status_col_var.get().strip().upper()}\n"
            f"Mode: {self.mode_var.get()}\n\n"
            "The selected values will be sent one-by-one to Windows Clipboard History.\n\n"
            "Each clipboard write is verified before the next value is sent. A slower delay is intentional for reliability.\n"
            "The spreadsheet will NOT be marked Completed automatically.\n\n"
            "Do you approve and start?"
        )
        if not answer:
            self.status_var.set("Task cancelled. Nothing was copied.")
            return

        self.values = selected
        self.populate_status(selected)
        self.stop_event.clear()
        self.processing = True
        self.load_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.complete_btn.config(state="disabled")
        self.progress_value.set(0)

        threading.Thread(
            target=self.worker, args=(selected, delay), daemon=True
        ).start()

    def update_row(self, index, clipboard_ok=True):
        iid, row_num, value, existing = self.status_rows[index]
        current = self.tree.item(iid, "values")
        sheet_status = current[3] if len(current) >= 4 else "Pending"
        self.tree.item(
            iid,
            values=(row_num, value, "✓ Copied" if clipboard_ok else "Not copied", sheet_status)
        )
        self.tree.see(iid)

    def worker(self, values, delay):
        total = len(values)
        done = 0
        stopped = False
        try:
            for i, (row_num, value, existing) in enumerate(values, start=1):
                if self.stop_event.is_set():
                    stopped = True
                    break

                set_clipboard_text(value)
                done = i
                self.after(0, self.update_row, i - 1, True)
                self.after(0, self.progress.configure, {"value": i * 100.0 / total})
                self.after(
                    0, self.status_var.set,
                    f"Copied {i:,} / {total:,}. Spreadsheet completion remains Pending."
                )

                if i < total and self.stop_event.wait(delay / 1000.0):
                    stopped = True
                    break

            if stopped:
                self.after(
                    0, self.status_var.set,
                    f"Stopped after {done:,} / {total:,}. Use Resume to continue later."
                )
            else:
                self.after(
                    0, self.status_var.set,
                    f"✓ Copy operation finished: {done:,} / {total:,}. Review the status table."
                )
                self.after(0, lambda: self.complete_btn.config(state="normal"))
                self.after(
                    0,
                    lambda: messagebox.showinfo(
                        "Copy Task Finished",
                        f"{done:,} values were copied to Windows Clipboard History.\n\n"
                        "The spreadsheet is still Pending.\n\n"
                        "Review the table, then approve completion if appropriate."
                    )
                )
        except Exception as e:
            self.after(
                0, self.status_var.set,
                f"Error after {done:,} / {total:,}: {e}. Spreadsheet remains unchanged."
            )
            self.after(0, lambda: messagebox.showerror(APP_TITLE, str(e)))
        finally:
            self.processing = False
            self.after(0, lambda: self.load_btn.config(state="normal"))
            self.after(0, lambda: self.stop_btn.config(state="disabled"))

    def request_completion(self):
        if self.processing:
            return

        total = len(self.status_rows)
        copied = sum(
            1 for iid, _, _, _ in self.status_rows
            if self.tree.item(iid, "values")[2] == "✓ Copied"
        )
        if total == 0:
            messagebox.showwarning(APP_TITLE, "There is no active task.")
            return
        if copied != total:
            messagebox.showwarning(
                "Cannot Complete",
                f"Only {copied} of {total} selected values are copied.\n\n"
                "The spreadsheet will not be marked Completed."
            )
            return

        data_col, status_col, start, end = self.get_range()

        answer = messagebox.askyesno(
            "Second Approval — Modify Spreadsheet",
            f"All {total:,} selected values are copied.\n\n"
            f"The app is now ready to write 'Completed' to column "
            f"{self.status_col_var.get().strip().upper()} for the selected rows.\n\n"
            "THIS WILL MODIFY YOUR SPREADSHEET FILE.\n\n"
            "Do you explicitly approve this spreadsheet update?"
        )
        if not answer:
            self.status_var.set("Completion not approved. Spreadsheet remains unchanged.")
            return

        try:
            self.write_status()
        except Exception as e:
            messagebox.showerror(
                APP_TITLE,
                "Copying is complete, but the spreadsheet status could not be updated:\n\n" + str(e)
            )
            self.status_var.set("Spreadsheet update failed. Range remains Pending.")
            return

        for iid, row_num, value, existing in self.status_rows:
            self.tree.item(
                iid, values=(row_num, value, "✓ Copied", "✓ Completed")
            )

        self.complete_btn.config(state="disabled")
        self.status_var.set(
            f"✓ RANGE COMPLETED — 'Completed' written to status column "
            f"{self.status_col_var.get().strip().upper()}."
        )

        self.set_summary(
            f"TASK SUMMARY\n"
            f"Rows selected: {self.from_var.get()} → {self.to_var.get()}\n"
            f"Values copied: {total:,}\n"
            f"Spreadsheet status: Completed\n"
            f"Status column updated: {self.status_col_var.get().strip().upper()}\n"
            f"Completion required explicit approval: YES"
        )

    def write_status(self):
        data_col, status_col, start, end = self.get_range()
        ext = os.path.splitext(self.file_path)[1].lower()

        if ext in (".xlsx", ".xlsm"):
            if openpyxl is None:
                raise RuntimeError("openpyxl is not installed.")

            sheet_name = self.sheet_var.get()
            self.close_workbook()

            keep_vba = ext == ".xlsm"
            wb = openpyxl.load_workbook(
                self.file_path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
                keep_links=False
            )
            try:
                ws = wb[sheet_name]
                task_rows = {row_num for _, row_num, _, _ in self.status_rows}
                for row_num in task_rows:
                    value = ws.cell(row=row_num, column=data_col).value
                    if value is not None and str(value).strip():
                        ws.cell(row=row_num, column=status_col).value = "Completed"
                wb.save(self.file_path)
            finally:
                wb.close()

            self.workbook = openpyxl.load_workbook(
                self.file_path, read_only=True, data_only=True, keep_links=False
            )
            self.sheet_names = self.workbook.sheetnames
            self.sheet_combo.config(state="readonly", values=self.sheet_names)
            self.sheet_var.set(sheet_name)
            return

        if ext == ".csv":
            with open(self.file_path, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))

            task_rows = {row_num for _, row_num, _, _ in self.status_rows}
            for row_num in sorted(task_rows):
                if row_num < 1 or row_num > len(rows):
                    continue
                row = rows[row_num - 1]
                data_value = row[data_col - 1] if data_col <= len(row) else ""
                if data_value is None or not str(data_value).strip():
                    continue
                while len(row) < status_col:
                    row.append("")
                row[status_col - 1] = "Completed"

            tmp = self.file_path + ".copy_paste_manager_tmp"
            try:
                with open(tmp, "w", encoding="utf-8-sig", newline="") as f:
                    csv.writer(f).writerows(rows)
                os.replace(tmp, self.file_path)
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            return

        raise ValueError("Unsupported file type.")

    def stop(self):
        if self.processing:
            self.stop_event.set()
            self.status_var.set("Stopping after the current clipboard operation…")

    def reset_status(self):
        if self.processing:
            messagebox.showwarning(APP_TITLE, "Stop the current task before resetting.")
            return
        self.values = []
        self.clear_tree()
        self.status_rows = []
        self.progress_value.set(0)
        self.complete_btn.config(state="disabled")
        self.set_summary("No active task.")
        self.status_var.set("Range reset. No task completion was saved by the app.")

    def on_close(self):
        self.stop_event.set()
        self.close_workbook()
        self.values = []
        self.status_rows = []
        self.destroy()


if __name__ == "__main__":
    CopyPasteManager().mainloop()
